import json
import os
import sys

from openpyxl import load_workbook, Workbook

def write_to_excel_file(file_name, sheet_name, data_to_write, row_number):
    wb = load_workbook(file_name) if os.path.exists(file_name) else Workbook()
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    for i in range(len(data_to_write)):
        ws.cell(row = row_number, column = i + 1).value = data_to_write[i]
    wb.save(file_name)

def find_inner_content(json_data):
    for inner_field, inner_value in json_data.items():
        print(inner_field, '\n\t', inner_value['name'], inner_value['id'])#, 
            #   '\n\t', inner_value['all_labels'] if 'all_labels' in json_data else 'ONLY ONE LABEL:', inner_value['label'], 
            #   '\n\t', inner_value['summary'], 
            #   '\n\t', inner_value['articles'],)
        if 'sub_topics' in inner_value:
            print('sub_topics:', '\n\t')
            inner_json_data = inner_value['sub_topics']
            find_inner_content(inner_json_data)

# from pathlib import Path
if __name__ == '__main__':
    input_file = sys.argv[1]
    print('Reading', input_file)
    # input_json_file = sys.argv[1]
    json_data = 0
    with open(input_file, 'r') as file_data:
        json_data = json.load(file_data)

    save_path = 'data/dfo/excel'
    os.makedirs(save_path, exist_ok = True)
    # excel_file_name = os.path.join(save_path, os.path.basename(input_file)[:-4]+'xlsx')

    find_inner_content(json_data)
        # # write_to_excel_file(excel_file_name, value['name'], mini_cluster_keywords, row)
        # if value['sub_topics']:
        #     inner_content = value['sub_topics']
        #     for inner_field, value in inner_content.items():
        #         print(field, '\n\t', value['name'], '\n\t', value['articles'][1,2])
        #         mini_cluster_keywords = [str(value['keywords'])]      
        
        