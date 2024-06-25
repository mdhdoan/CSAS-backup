import json
import os
import sys

from openpyxl import load_workbook, Workbook

def write_to_excel_file(file_name, sheet_name, data_to_write, row_number):
    wb = load_workbook(file_name) if os.path.exists(file_name) else Workbook()
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    for i in range(len(data_to_write)):
        ws.cell(row = row_number, column = i + 1).value = data_to_write[i]
    wb.save(file_name)

# from pathlib import Path
if __name__ == '__main__':
    input_directory = sys.argv[1]
    for file in os.listdir(input_directory):
        print('Reading', file)
        file_name = os.path.join(input_directory, file)
        if not file_name.endswith('.json') or file_name.endswith('--1-cls.json'):
            continue
        else:
            input_json_file = file
            # input_json_file = sys.argv[1]
            json_data = 0
            with open(file_name, 'r') as file_data:
                json_data = json.load(file_data)

            save_path = 'data/dfo/excel'
            os.makedirs(save_path, exist_ok = True)
            excel_file_name = os.path.join(save_path, os.path.basename(input_json_file)[:-4]+'xlsx')

            row = 1
            for field, value in json_data.items():
                # print(field, '\n\t', value['name'], '\n\t', value['keywords'])
                mini_cluster_keywords = [str(value['keywords'])]
                write_to_excel_file(excel_file_name, value['name'],
                                        mini_cluster_keywords, row)
                row += 1
                for article_id, article_title, article_url, article_cluster_score in value['articles']:
                    # print('\n\t', article_id, article_title, article_url)
                    writing_data = [article_id, article_title, article_url]
                    write_to_excel_file(excel_file_name, value['name'],
                                        writing_data, row)
                    row += 1
                row = 1
        