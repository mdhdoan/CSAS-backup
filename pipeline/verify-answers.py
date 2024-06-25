import sys

from openpyxl import load_workbook, Workbook

from langchain.output_parsers.list import NumberedListOutputParser
from langchain.prompts.prompt import PromptTemplate

from langchain_community.llms import Ollama

from sentence_transformers import SentenceTransformer, util


def load_data_file(file_name, sheet_name):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]
    headers, inputs = [], []
    for i, row in enumerate(ws): 
        if i == 0:
            headers = [cell.value for cell in row]
            continue
        inputs.append({headers[j]: cell.value for j, cell in enumerate(row)})
    return inputs
 

def create_generating_question_prompt(format_instructions):
    QA_TEMPLATE = """Use the text below delimited by triple quotes (```) to create a GENERIC question that can be used later for similar texts. Return ONLY the question WITHOUT any explanation.
    TEXT:```{text}```

    QUESTION:{format_instructions}:"""
    return PromptTemplate(
        input_variables=["text"], 
        partial_variables={"format_instructions": format_instructions}, 
        template=QA_TEMPLATE)


def create_extracting_answer_prompt(format_instructions):
    QA_TEMPLATE = """Give a meaningful answer to the question using the text below delimited by triple quotes (```). Return ONLY the answer WITHOUT any explanation.
    QUESTION: {question}
    
    TEXT:```{text}```

    ANSWER:{format_instructions}:"""
    return PromptTemplate(
        input_variables=["question", "text"], 
        partial_variables={"format_instructions": format_instructions}, 
        template=QA_TEMPLATE)
    
    
def write_to_excel_file(file_name, sheet_name, outputs):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(outputs[0].keys())
    ws.append(headers)
    for output in outputs:
        row = [output[h] for h in headers]
        ws.append(row)
    wb.save(file_name)


if __name__ == '__main__':
    input_data_file = sys.argv[1]
    input_sheet_name = sys.argv[2]
    output_sheet_name = sys.argv[3]
    inputs = load_data_file(input_data_file, input_sheet_name)
    
    output_parser = NumberedListOutputParser()
    format_instructions = output_parser.get_format_instructions()

    llm_model_name = 'llama3'
    llm_model = Ollama(model=llm_model_name, temperature=0.0)

    generating_question_prompt = create_generating_question_prompt(format_instructions)
    gq_llm_chain = generating_question_prompt | llm_model | output_parser

    extracting_answer_prompt = create_extracting_answer_prompt(format_instructions)
    ea_llm_chain = extracting_answer_prompt | llm_model | output_parser
    
    st_model_name = "all-mpnet-base-v2"
    sentence_transformer = SentenceTransformer(st_model_name)

    outputs = []
    for input in inputs:
        result = gq_llm_chain.invoke({"text": input['Answer']})
        output = {
            'status': 'OK' if result else 'ERR',
            'similarity': -1,
            'generated question': result[0] if result else '',
            'extracted answer': None,
            'original question': input['Question'],
            'original answer': input['Answer'],
            'url': input['URL'],
            'title': input['Title'],
        }
        
        if not result:
            outputs.append(output)
            print('\t'.join([f"{v}" for _, v in output.items()]))
            continue

        result = ea_llm_chain.invoke({"question": output['generated question'], "text": input['Full text']})
        if not result:
            output['status'] = 'ERR'
            output['extracted answer'] = None
            outputs.append(output)
            print('\t'.join([f"{v}" for _, v in output.items()]))
            continue
        
        output['extracted answer'] = result[0]
        original_embedding = sentence_transformer.encode(output['original answer'])
        extracted_embedding = sentence_transformer.encode(output['extracted answer'])
        output['similarity'] = float(util.cos_sim(original_embedding, extracted_embedding))
        
        if output['similarity'] < 0.9:
            output['status'] = 'LOW'

        outputs.append(output)
        print('\t'.join([f"{v}" for _, v in output.items()]))

    output_excel_file = input_data_file.replace(".xlsx", f"-{output_sheet_name}.xlsx")
    write_to_excel_file(output_excel_file, output_sheet_name, outputs)
