import sys

from openpyxl import load_workbook

from langchain.output_parsers.list import NumberedListOutputParser
from langchain.prompts.prompt import PromptTemplate

from langchain_community.llms import Ollama

def load_data_file(file_name, sheet_name):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]
    headers, inputs = [], []
    for i, row in enumerate(ws): 
        if i == 0:
            headers = [cell.value for cell in row]
            continue
        inputs.append({headers[j]: cell.value for j, cell in enumerate(row)})
    return inputs
 

def create_qa_prompt(format_instructions):
    QA_TEMPLATE = """Use the text below delimited by triple quotes (```) to create a GENERIC question that can be used later for similar texts. Return ONLY the question WITHOUT any explanation.
    TEXT:```{text}```

    QUESTION:{format_instructions}:"""
    return PromptTemplate(
        input_variables=["text"], 
        partial_variables={"format_instructions": format_instructions}, 
        template=QA_TEMPLATE)
    
def write_to_excel_file(file_name, sheet_name, data_to_write, row_number):
    wb = load_workbook(file_name)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    for i in range(len(data_to_write)):
        ws.cell(row = row_number, column = i + 1).value = data_to_write[i]
    wb.save(file_name)


### command: python [script] [excel file] [excel sheet]
if __name__ == '__main__':
    input_data_file = sys.argv[1]
    sheet_name = sys.argv[2]
    inputs = load_data_file(input_data_file, sheet_name)
    
    output_parser = NumberedListOutputParser()
    format_instructions = output_parser.get_format_instructions()

    model_name = 'llama3'
    llm_model = Ollama(model=model_name, temperature=0.0)

    prompt = create_qa_prompt(format_instructions)
    llm_chain = prompt | llm_model | output_parser

    results = llm_chain.batch([{"text": input['Answer']} for input in inputs], config={"max_concurrency": 8})
    for i, result in enumerate(results):
        if not result:
            print(f"ERROR\tNone\t{inputs[i]['Question']}\t{inputs[i]['Answer']}")
            status = 'ERROR'
            generated_question = 'None'
        else:
            print(f"OK\t{result[0]}\t{inputs[i]['Question']}\t{inputs[i]['Answer']}")
            status = 'OK'
            generated_question = result[0]
        data_to_write = [status, generated_question, inputs[i]['Question'], inputs[i]['Answer']]
        row_number = i + 1
        write_to_excel_file(input_data_file, 'generated', data_to_write, row_number)
    